import csv
import math
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from jinja2 import Environment, FileSystemLoader
import pathlib
import pdfkit


def csv_reader():
    with open(input('Введите название файла: '), encoding='utf-8-sig') as r_file:
        file = csv.reader(r_file, delimiter=",")
        count = 0
        data = []
        name = []
        for line in file:
            if count == 0:
                lenLine = len(line)
                count += 1
                name = line
            else:
                if len(line) != lenLine or '' in line:
                    continue
                data.append(line)
        return data, name

data, name = csv_reader()
profession = input('Введите название профессии: ')

type_of_file = input('Введите данные для печати:')

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}


def get_vacs(data, names):
    vacancies = []
    for vacancy in data:
        name = vacancy[names.index('name')]
        salary_currency = vacancy[names.index('salary_currency')]
        salary_from = int(float(vacancy[names.index('salary_from')]))*currency_to_rub[salary_currency]
        salary_to = int(float(vacancy[names.index('salary_to')]))*currency_to_rub[salary_currency]
        area_name = vacancy[names.index('area_name')]
        published_at = vacancy[names.index('published_at')]
        vacancies.append([name, salary_from, salary_to, salary_currency, area_name, published_at])

    return vacancies


def get_dynamic(vacancies):
    salaries = {}
    vacs_count = {}
    vac_count_prof = {}
    salaries_prof = {}
    salaries_city = {}
    count_city = {}
    for line in vacancies:
        year = int(line[-1][0:4])
        city = line[-2]
        salary = (line[1] + line[2]) / 2
        if year not in salaries.keys():
            salaries[year] = []
            salaries[year].append(salary)
        else:
            salaries[year].append(salary)

        if year not in vacs_count.keys():
            vacs_count[year] = 1
        else:
            vacs_count[year] += 1

        if profession in line[0]:
            if year not in salaries_prof.keys():
                salaries_prof[year] = []
                salaries_prof[year].append(salary)
            else:
                salaries_prof[year].append(salary)

            if year not in vac_count_prof.keys():
                vac_count_prof[year] = 1
            else:
                vac_count_prof[year] += 1

        if city in salaries_city:
            salaries_city[city].append(salary)
        else:
            salaries_city[city] = []
            salaries_city[city].append(salary)

        if city in count_city:
            count_city[city] += 1
        else:
            count_city[city] = 1

    dynamic_salary = {}
    dynamic_salary_prof = {}
    salary_level_city = {}
    vacancies_city = {}

    for key in salaries:
        dynamic_salary[key] = math.floor(sum(salaries[key]) / len(salaries[key]))

    for key in salaries_prof:
        dynamic_salary_prof[key] = math.floor(sum(salaries_prof[key]) / len(salaries_prof[key]))

    for key in count_city:
        vacancies_city[key] = round(count_city[key] / len(vacancies), 4)

    for key in salaries_city:
        if vacancies_city[key] >= 0.01:
            salary_level_city[key] = math.floor(sum(salaries_city[key]) / len(salaries_city[key]))

    if len(dynamic_salary_prof.keys()) == 0:
        dynamic_salary_prof = {2022: 0}

    if len(vac_count_prof.keys()) == 0:
        vac_count_prof = {2022: 0}

    return dynamic_salary, vacs_count, dynamic_salary_prof, vac_count_prof, salary_level_city, vacancies_city


vacancies = get_vacs(data, name)
dynamic_salary, dynamic_vac_count, dynamic_salary_profession, dynamic_vac_profession, salary_level_city, vacancies_city = get_dynamic(vacancies)

sort_salary_level_city = sorted(salary_level_city.items(), key=lambda item: item[1], reverse=True)
sort_salary_level_city = dict(sort_salary_level_city)
sort_vacancies_city = sorted(vacancies_city.items(), key=lambda item: item[1], reverse=True)
sort_vacancies_city = dict(sort_vacancies_city)

sort_salary_level_city_cor = {}
sort_vacancies_city_cor = {}

count = 0
for key in sort_salary_level_city:
    count += 1
    sort_salary_level_city_cor[key] = sort_salary_level_city[key]
    if count == 10:
        break

count = 0
for key in sort_vacancies_city:
    count += 1
    if vacancies_city[key] >= 0.01:
        sort_vacancies_city_cor[key] = sort_vacancies_city[key]
    else:
        break
    if count == 10:
        break

print(f'Динамика уровня зарплат по годам: {dynamic_salary}')
print(f'Динамика количества вакансий по годам: {dynamic_vac_count}')
print(f'Динамика уровня зарплат по годам для выбранной профессии: {dynamic_salary_profession}')
print(f'Динамика количества вакансий по годам для выбранной профессии: {dynamic_vac_profession}')
print(f'Уровень зарплат по городам (в порядке убывания): {sort_salary_level_city_cor}')
print(f'Доля вакансий по городам (в порядке убывания): {sort_vacancies_city_cor}')


labels_years = list(dynamic_salary.keys())
mid_salary = list(dynamic_salary.values())
mid_salary_profession = list(dynamic_salary_profession.values())
vac_count = list(dynamic_vac_count.values())
vac_count_profession = list(dynamic_vac_profession.values())
cities = list(sort_salary_level_city_cor.keys())
labels_parts = list(sort_vacancies_city_cor.keys())
salaries_parts = list(sort_vacancies_city_cor.values())
others = 1 - sum(salaries_parts)
labels_parts.append('Другие')
salaries_parts.append(others)

for i in range(0, len(cities)):
    cities[i] = cities[i].replace(' ', '\n')
    if '-' in cities[i]:
        ind = cities[i].index('-')
        cities[i] = cities[i][0:ind+1] + '\n' + cities[i][ind+1:]

salary_level = list(sort_salary_level_city_cor.values())

x = np.arange(len(labels_years))
width = 0.4

fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)

# первый график
rects1 = ax1.bar(x - width/2, mid_salary, width, label='средняя з/п')
rects2 = ax1.bar(x + width/2, mid_salary_profession, width, label=f'з/п {profession}')
ax1.set_xticks(x, labels_years)
ax1.set_xticklabels(labels_years, rotation=90, fontsize=8)
yticks = []
for i in ax1.get_yticks():
    yticks.append(int(i))
ax1.set_yticklabels(yticks, fontsize=8)
ax1.legend(fontsize=8)
ax1.grid(axis='y')
ax1.set_title('Уровень зарплат по годам', fontsize=8)

#второй график
rects3 = ax2.bar(x - width/2, vac_count, width, label='Количество вакансий')
rects4 = ax2.bar(x + width/2, vac_count_profession, width, label=f'Количество вакансий \n{profession}')
ax2.set_xticks(x, labels_years)
ax2.set_xticklabels(labels_years, rotation=90, fontsize=8)
yticks = []
for i in ax2.get_yticks():
    yticks.append(int(i))
ax2.set_yticklabels(yticks, fontsize=8)
ax2.legend(fontsize=8)
ax2.grid(axis='y')
ax2.set_title('Количество вакансий по годам', fontsize=8)

#третий график
y_pos = np.arange(len(cities))
ax3.barh(y_pos, salary_level, align='center')
ax3.set_yticks(y_pos, cities, fontsize=6)
ax3.grid(axis='x')
xticks = []
for i in ax3.get_xticks():
    xticks.append(int(i))
ax3.set_xticklabels(xticks, fontsize=8)
ax3.set_title('Уровень зарплат по городам', fontsize=8)

#четвертый график
ax4.pie(salaries_parts, labels=labels_parts, textprops={'fontsize': 6})
ax4.set_title('Доля вакансий по городам', fontsize=8)

fig.tight_layout()
plt.savefig('graph.png')

wb = Workbook()
ws = wb.active
side = Side(border_style='thin', color='000000')
ws.title = 'Статистика по годам'


ws.append({'A': 'Год', 'B': 'Средняя зарплата', 'C': f'Средняя зарплата - {profession}', 'D': 'Количество ваканский',
           'E': f'Количество ваканский - {profession}'})

ws.column_dimensions['A'].w = 5
ws.column_dimensions['B'].w = 17
ws.column_dimensions['C'].w = 20 + len(profession)
ws.column_dimensions['D'].w = 21
ws.column_dimensions['E'].w = 24 + len(profession)

ws['A1'].font = Font(bold=True)
ws['A1'].border = Border(top=side, left=side, bottom=side, right=side)
ws['B1'].font = Font(bold=True)
ws['B1'].border = Border(top=side, left=side, bottom=side, right=side)
ws['C1'].font = Font(bold=True)
ws['C1'].border = Border(top=side, left=side, bottom=side, right=side)
ws['D1'].font = Font(bold=True)
ws['D1'].border = Border(top=side, left=side, bottom=side, right=side)
ws['E1'].font = Font(bold=True)
ws['E1'].border = Border(top=side, left=side, bottom=side, right=side)


for count, key in enumerate(dynamic_salary):
    ws.append([key, dynamic_salary[key], dynamic_salary_profession[key], dynamic_vac_count[key], dynamic_vac_profession[key]])
    ws[f'A{count+2}'].border = Border(top=side, left=side, bottom=side, right=side)
    ws[f'B{count+2}'].border = Border(top=side, left=side, bottom=side, right=side)
    ws[f'C{count+2}'].border = Border(top=side, left=side, bottom=side, right=side)
    ws[f'D{count+2}'].border = Border(top=side, left=side, bottom=side, right=side)
    ws[f'E{count+2}'].border = Border(top=side, left=side, bottom=side, right=side)


ws1 = wb.create_sheet()
ws1.title = 'Статистика по городам'

ws1.append({'A': 'Город', 'B': 'Уровень зарплат', 'D': 'Город', 'E': 'Доля вакансий'})
ws1.column_dimensions['C'].w = 1.29
ws1.column_dimensions['A'].w = 20
ws1.column_dimensions['B'].w = 16
ws1.column_dimensions['D'].w = 20
ws1.column_dimensions['E'].w = 14

ws1['A1'].font = Font(bold=True)
ws1['A1'].border = Border(top=side, left=side, bottom=side, right=side)
ws1['B1'].font = Font(bold=True)
ws1['B1'].border = Border(top=side, left=side, bottom=side, right=side)
ws1['D1'].font = Font(bold=True)
ws1['D1'].border = Border(top=side, left=side, bottom=side, right=side)
ws1['E1'].font = Font(bold=True)
ws1['E1'].border = Border(top=side, left=side, bottom=side, right=side)

maxlen = 0
for count, key in enumerate(sort_salary_level_city_cor):
    lenkey = len(key)
    if lenkey > maxlen:
        maxlen = lenkey
    ws1.append([key, sort_salary_level_city_cor[key]])
    ws1[f'A{count + 2}'].border = Border(top=side, left=side, bottom=side, right=side)
    ws1[f'B{count + 2}'].border = Border(top=side, left=side, bottom=side, right=side)
ws1.column_dimensions['A'].w = maxlen + 2

maxlen = 0
for count, key in enumerate(sort_vacancies_city_cor):
    lenkey = len(key)
    if lenkey > maxlen:
        maxlen = lenkey
    ws1[f'D{count + 2}'] = key
    ws1[f'E{count + 2}'] = sort_vacancies_city_cor[key]
    ws1[f'E{count + 2}'].number_format = BUILTIN_FORMATS[10]
    ws1[f'D{count + 2}'].border = Border(top=side, left=side, bottom=side, right=side)
    ws1[f'E{count + 2}'].border = Border(top=side, left=side, bottom=side, right=side)
ws1.column_dimensions['D'].w = maxlen + 2

wb.save('report.xlsx')


env = Environment(loader=FileSystemLoader('../templates'))
template = env.get_template("pdf_template.html")
stats = []
for year in dynamic_salary.keys():
    stats.append([year, dynamic_salary[year], dynamic_vac_count[year], dynamic_salary_profession[year], dynamic_vac_profession[year]])

for key in sort_vacancies_city_cor:
    sort_vacancies_city_cor[key] = round(sort_vacancies_city_cor[key] * 100, 2)

pdf_template = template.render(
    {'name': profession, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
     'stats': stats, 'stats5': sort_salary_level_city_cor, 'stats6': sort_vacancies_city_cor})

config = pdfkit.configuration(wkhtmltopdf=r'/usr/bin/wkhtmltopdf')
pdfkit.from_string(pdf_template, 'temp.pdf', configuration=config, options={"enable-local-file-access": ""})
pdfkit.from_string(pdf_template, 'temp.pdf', options={"enable-local-file-access": ""})